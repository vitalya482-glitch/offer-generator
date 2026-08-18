from __future__ import annotations

from pathlib import Path
import re
import sys

from openpyxl import load_workbook
from PySide6.QtWidgets import QComboBox, QGridLayout

from core.excel_reader import parse_stulz_calc


_SUPPORTED_CURRENCIES = {"KZT", "EUR", "USD"}


def _plain(value: object) -> str:
    return str(value or "").replace("\xa0", " ").strip()


def _currency_from_text(value: object) -> str:
    raw = _plain(value)
    if not raw:
        return ""
    text = raw.lower().replace("ё", "е")
    text = re.sub(r"[\r\n\t]+", " ", text)

    if "₸" in raw or re.search(r"(?:^|[^a-zа-я0-9])(kzt|тенге|тг)(?:$|[^a-zа-я0-9])", text):
        return "KZT"
    if "€" in raw or re.search(r"(?:^|[^a-zа-я0-9])(eur|euro|евро)(?:$|[^a-zа-я0-9])", text):
        return "EUR"
    if "$" in raw or re.search(r"(?:^|[^a-zа-я0-9])(usd|доллар(?:ов|а|ы)?)(?:$|[^a-zа-я0-9])", text):
        return "USD"
    return ""


def detect_stulz_currency(calc_path: str | Path, sheet_name: str | None) -> str:
    """Conservatively detect the customer currency of a STULZ Excel calculation.

    Unlike the legacy STULZ parser this helper never assumes EUR merely because
    no currency marker was found. A non-trivial exchange rate is a strong KZT
    signal in the SAM calculation templates. Otherwise explicit currency text or
    Excel number formats are used. Ambiguous sheets return an empty string so the
    user must confirm the currency manually before generating the offer.
    """

    path = Path(calc_path)
    if not path.exists() or path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return ""

    try:
        calc = parse_stulz_calc(path, sheet_name or None)
        rate = float(getattr(calc, "exchange_rate", 1) or 1)
        if rate > 1.01:
            return "KZT"
    except Exception:
        pass

    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        if sheet_name and sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
        else:
            ws = workbook[workbook.sheetnames[0]]

        # Prefer an explicit marker in the selected sheet title.
        title_currency = _currency_from_text(ws.title)
        if title_currency:
            return title_currency

        found: set[str] = set()
        max_row = min(int(ws.max_row or 0), 140)
        max_col = min(int(ws.max_column or 0), 100)
        if max_row <= 0 or max_col <= 0:
            return ""

        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                for source in (getattr(cell, "value", None), getattr(cell, "number_format", "")):
                    currency = _currency_from_text(source)
                    if currency:
                        found.add(currency)

        if len(found) == 1:
            return next(iter(found))
        return ""
    except Exception:
        return ""
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass


def _find_files_grid(widget: object) -> QGridLayout | None:
    client_edit = getattr(widget, "client_edit", None)
    card = client_edit.parentWidget() if client_edit is not None else None
    layout = card.layout() if card is not None else None
    if layout is None:
        return None

    for index in range(layout.count()):
        child = layout.itemAt(index).layout()
        if isinstance(child, QGridLayout):
            return child
    return None


def _move_grid_row(grid: QGridLayout, old_row: int, new_row: int) -> None:
    for col in range(3):
        item = grid.itemAtPosition(old_row, col)
        if item is None:
            continue
        widget = item.widget()
        if widget is None:
            continue
        grid.removeWidget(widget)
        grid.addWidget(widget, new_row, col)


def _currency_source_key(self) -> str:
    try:
        calc_path = self._path_from_combo(self.calc_combo).strip()
    except Exception:
        calc_path = ""
    sheet = self.sheet_combo.currentText().strip() if hasattr(self, "sheet_combo") else ""
    return f"{calc_path}|{sheet}"


def currency_value(self) -> str:
    combo = getattr(self, "currency_combo", None)
    if combo is None:
        return ""
    return str(combo.currentData() or "").upper().strip()


def _set_currency_value(self, value: str) -> None:
    combo = getattr(self, "currency_combo", None)
    if combo is None:
        return
    value = (value or "").upper().strip()
    setattr(self, "_stulz_setting_currency_programmatically", True)
    try:
        for index in range(combo.count()):
            if str(combo.itemData(index) or "").upper().strip() == value:
                combo.setCurrentIndex(index)
                return
        combo.setCurrentIndex(0)
    finally:
        setattr(self, "_stulz_setting_currency_programmatically", False)


def _auto_detect_currency(self, force: bool = False) -> None:
    combo = getattr(self, "currency_combo", None)
    if combo is None:
        return

    source_key = _currency_source_key(self)
    previous_key = str(getattr(self, "_stulz_currency_source_key", "") or "")
    if not force and source_key == previous_key and currency_value(self):
        return

    try:
        calc_path = self._path_from_combo(self.calc_combo).strip()
    except Exception:
        calc_path = ""
    sheet_name = self.sheet_combo.currentText().strip() if hasattr(self, "sheet_combo") else ""
    detected = detect_stulz_currency(calc_path, sheet_name)

    _set_currency_value(self, detected)
    self._stulz_currency_source_key = source_key
    try:
        self.settings.setValue("stulz_currency", detected)
        self.settings.setValue("stulz_currency_source_key", source_key)
        self.settings.sync()
    except Exception:
        pass


def _on_currency_changed(self, *_args) -> None:
    if getattr(self, "_stulz_setting_currency_programmatically", False):
        return
    source_key = _currency_source_key(self)
    self._stulz_currency_source_key = source_key
    try:
        self.settings.setValue("stulz_currency", currency_value(self))
        self.settings.setValue("stulz_currency_source_key", source_key)
        self.settings.sync()
    except Exception:
        pass
    try:
        self.refresh_preview()
    except Exception:
        pass


def _ensure_currency_control(self) -> None:
    if getattr(self, "_stulz_currency_control_installed", False):
        return

    grid = _find_files_grid(self)
    if grid is None:
        return

    # Existing rows are: 0 client, 1 Excel, 2 sheet, 3 Word, 4 specs, 5 output.
    # Move the lower rows down and insert Currency directly after the Excel sheet.
    for row in (5, 4, 3):
        _move_grid_row(grid, row, row + 1)

    combo = QComboBox()
    combo.addItem("Не указана", "")
    combo.addItem("KZT", "KZT")
    combo.addItem("EUR", "EUR")
    combo.addItem("USD", "USD")
    combo.setToolTip(
        "Валюта коммерческого предложения. Определяется автоматически из выбранного Calc; "
        "если определить нельзя — выберите вручную."
    )
    self.owner._add_row(grid, 3, "Валюта", combo, None, None)
    combo.currentIndexChanged.connect(self._on_stulz_currency_changed)

    self.currency_combo = combo
    self._stulz_currency_control_installed = True

    current_key = _currency_source_key(self)
    saved_key = str(self.settings.value("stulz_currency_source_key", "") or "")
    saved_currency = str(self.settings.value("stulz_currency", "") or "").upper().strip()
    if current_key and current_key == saved_key and saved_currency in _SUPPORTED_CURRENCIES:
        _set_currency_value(self, saved_currency)
        self._stulz_currency_source_key = current_key
    else:
        _auto_detect_currency(self, force=True)


def make_context(self):
    context = self._stulz_currency_original_make_context()
    options = dict(getattr(context, "brand_options", None) or {})
    options["stulz_currency"] = currency_value(self)
    context.brand_options = options
    return context


def validate_context(self, context) -> None:
    self._stulz_currency_original_validate_context(context)
    value = str((getattr(context, "brand_options", None) or {}).get("stulz_currency", "") or "").upper().strip()
    if value not in _SUPPORTED_CURRENCIES:
        raise ValueError(
            "Валюта не определена. Проверьте поле «Валюта» и выберите KZT, EUR или USD перед формированием КП."
        )


def remember_values(self) -> None:
    self._stulz_currency_original_remember_values()
    try:
        self.settings.setValue("stulz_currency", currency_value(self))
        self.settings.setValue("stulz_currency_source_key", _currency_source_key(self))
        self.settings.sync()
    except Exception:
        pass


def clear_cache(self) -> None:
    self._stulz_currency_original_clear_cache()
    try:
        self.settings.remove("stulz_currency")
        self.settings.remove("stulz_currency_source_key")
        self.settings.sync()
    except Exception:
        pass
    _set_currency_value(self, "")
    self._stulz_currency_source_key = ""


def scan_project(self, force: bool = False) -> None:
    old_key = _currency_source_key(self)
    self._stulz_currency_original_scan_project(force)
    new_key = _currency_source_key(self)
    if new_key != old_key:
        _auto_detect_currency(self, force=True)


def load_sheets(self, refresh: bool = True) -> None:
    old_key = _currency_source_key(self)
    self._stulz_currency_original_load_sheets(refresh)
    new_key = _currency_source_key(self)
    if new_key != old_key:
        _auto_detect_currency(self, force=True)


def _install() -> None:
    page_module = sys.modules.get("gui.pages.stulz_page")
    page_class = getattr(page_module, "StulzPage", None) if page_module is not None else None
    if page_class is None:
        return

    originals = {
        "make_context": "_stulz_currency_original_make_context",
        "validate_context": "_stulz_currency_original_validate_context",
        "remember_values": "_stulz_currency_original_remember_values",
        "clear_cache": "_stulz_currency_original_clear_cache",
        "scan_project": "_stulz_currency_original_scan_project",
        "load_sheets": "_stulz_currency_original_load_sheets",
    }
    for method_name, saved_name in originals.items():
        if not hasattr(page_class, saved_name):
            setattr(page_class, saved_name, getattr(page_class, method_name))

    for name, function in {
        "_currency_source_key": _currency_source_key,
        "currency_value": currency_value,
        "_set_currency_value": _set_currency_value,
        "_auto_detect_currency": _auto_detect_currency,
        "_on_stulz_currency_changed": _on_currency_changed,
        "_ensure_currency_control": _ensure_currency_control,
        "make_context": make_context,
        "validate_context": validate_context,
        "remember_values": remember_values,
        "clear_cache": clear_cache,
        "scan_project": scan_project,
        "load_sheets": load_sheets,
    }.items():
        setattr(page_class, name, function)

    try:
        from PySide6.QtWidgets import QApplication
        app = QApplication.instance()
        if app is not None:
            for widget in app.allWidgets():
                if isinstance(widget, page_class):
                    widget._ensure_currency_control()
                    widget._auto_detect_currency(force=False)
                    widget.refresh_preview()
    except Exception:
        pass


_install()
