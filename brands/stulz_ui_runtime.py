from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from PySide6.QtCore import Qt, QTimer
from PySide6.QtWidgets import QApplication, QLabel

import brands.stulz_runtime as _runtime
from core.excel_reader import CachedSheet, parse_model_groups


for _name in dir(_runtime):
    if not _name.startswith("__"):
        globals()[_name] = getattr(_runtime, _name)


_SPEC_PREFIXES = (
    "Спецификации:",
    "Чертежи для вставки:",
    "Опций для спецификации:",
    "Строк тех. характеристик:",
    "Модели:",
    "Количество:",
    "Файлы чертежей:",
)

_FINANCE_MARKERS = (
    "financing",
    "finance cost",
    "financial cost",
    "финансирован",
)


def _number(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace("\u00a0", " ").replace(" ", "").replace(",", ".").replace("%", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except Exception:
        return None


def _read_financing(calc_path: Path, sheet_name: str | None) -> dict[str, Any]:
    """Read already-calculated financing values for GUI information only."""

    result: dict[str, Any] = {"found": False, "percent": None, "amount": None}
    if not calc_path.exists():
        return result

    try:
        workbook = load_workbook(calc_path, read_only=True, data_only=True)
        raw_sheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        sheet = CachedSheet(raw_sheet)
        groups = parse_model_groups(sheet)
        amount_cols = [amount_col for _qty_col, amount_col, _model in groups]

        percents: list[float] = []
        amounts: list[float] = []

        for row, col, normalized in sheet.norm_cells():
            if not any(marker in normalized for marker in _FINANCE_MARKERS):
                continue

            result["found"] = True
            label = str(sheet.cell(row, col).value or "")
            label_norm = normalized

            percent_match = re.search(r"(\d+(?:[.,]\d+)?)\s*%", label)
            if percent_match:
                percents.append(float(percent_match.group(1).replace(",", ".")))

            row_numbers: list[tuple[int, float]] = []
            for current_col in range(1, sheet.max_column + 1):
                if current_col == col:
                    continue
                value = _number(sheet.cell(row, current_col).value)
                if value is not None and abs(value) > 1e-12:
                    row_numbers.append((current_col, value))

            percent_row = "%" in label or "percent" in label_norm or "процент" in label_norm
            amount_row = "amount" in label_norm or "сумм" in label_norm or "стоим" in label_norm

            group_amounts: list[float] = []
            for amount_col in amount_cols:
                value = _number(sheet.cell(row, amount_col).value)
                if value is not None and abs(value) > 100:
                    group_amounts.append(value)

            if group_amounts and not percent_row:
                amounts.append(sum(group_amounts))
            elif amount_row:
                large = [value for _col, value in row_numbers if abs(value) > 100]
                if large:
                    amounts.append(sum(large))

            for current_col, value in row_numbers:
                if abs(current_col - col) > 4 and not percent_row:
                    continue
                if 0 < abs(value) <= 1:
                    percents.append(abs(value) * 100)
                elif 1 < abs(value) <= 100:
                    percents.append(abs(value))

        if percents:
            result["percent"] = percents[0]
        if amounts:
            result["amount"] = max(amounts, key=abs)

        try:
            workbook.close()
        except Exception:
            pass
    except Exception:
        pass

    return result


def _partition_preview(raw_text: str) -> tuple[list[str], list[str], str | None]:
    """Move specification-derived lines right and total to the bottom."""

    left: list[str] = []
    spec: list[str] = []
    total: str | None = None
    warning_rows = False

    for line in raw_text.splitlines():
        stripped = line.strip()

        if stripped.startswith("Сумма:"):
            total = "Итоговая сумма:" + stripped[len("Сумма:") :]
            warning_rows = False
            continue

        if any(stripped.startswith(prefix) for prefix in _SPEC_PREFIXES):
            spec.append(stripped)
            warning_rows = False
            continue

        if stripped == "Предупреждения:":
            spec.append(stripped)
            warning_rows = True
            continue

        if stripped.startswith("Предупреждения:"):
            spec.append(stripped)
            warning_rows = False
            continue

        if warning_rows:
            if stripped.startswith("- "):
                spec.append(stripped)
                continue
            if not stripped:
                warning_rows = False
                continue
            warning_rows = False

        left.append(line)

    compact: list[str] = []
    previous_blank = False
    for line in left:
        blank = not line.strip()
        if blank and previous_blank:
            continue
        compact.append(line)
        previous_blank = blank

    while compact and not compact[-1].strip():
        compact.pop()

    return compact, spec, total


def _insert_after(lines: list[str], prefix: str, new_lines: list[str]) -> list[str]:
    if not new_lines:
        return lines
    result = list(lines)
    index = next((i + 1 for i, line in enumerate(result) if line.strip().startswith(prefix)), len(result))
    result[index:index] = new_lines
    return result


def _ensure_spec_summary_widget(self) -> None:
    if getattr(self, "_stulz_spec_summary_installed", False):
        return

    parent = self.spec_models_table.parentWidget()
    layout = parent.layout() if parent is not None else None
    if layout is None:
        return

    summary = QLabel("Спецификации пока не прочитаны.")
    summary.setWordWrap(True)
    summary.setTextInteractionFlags(Qt.TextSelectableByMouse)
    summary.setObjectName("Hint")

    table_index = layout.indexOf(self.spec_models_table)
    if table_index >= 0:
        layout.insertWidget(table_index, summary)
    else:
        layout.addWidget(summary)

    self._stulz_spec_summary_label = summary
    self._stulz_spec_summary_installed = True


def _set_spec_summary(self, lines: list[str] | str) -> None:
    self._ensure_spec_summary_widget()
    label = getattr(self, "_stulz_spec_summary_label", None)
    if label is None:
        return
    text = lines if isinstance(lines, str) else "\n".join(line for line in lines if line.strip())
    label.setText(text or "Спецификации пока не прочитаны.")


def refresh_preview(self) -> None:
    """STULZ GUI preview split into calculation info and specification info."""

    self._ensure_spec_summary_widget()

    try:
        context = self.make_context()
        self.refresh_spec_models(context)

        if not context.calc_path.exists():
            self.preview.setPlainText("Excel-файл пока не выбран или не найден.")
            self._set_spec_summary("Спецификации пока не прочитаны.")
            return

        raw_text = _runtime.preview(context)
        left_lines, spec_lines, total_line = _partition_preview(raw_text)

        calc = _runtime.load_calc(context)
        financing = _read_financing(context.calc_path, calc.sheet_name)
        finance_lines: list[str] = []

        if financing.get("found"):
            percent = financing.get("percent")
            amount = financing.get("amount")

            if percent is not None:
                finance_lines.append(f"Финансирование: {_runtime.format_qty(float(percent))}%")
            elif amount is None:
                finance_lines.append("Финансирование: включено")

            if amount is not None:
                finance_lines.append(
                    f"Сумма финансирования: {_runtime.format_money(float(amount))} "
                    f"{_runtime.currency_name(calc.currency)}"
                )

        left_lines = _insert_after(left_lines, "Монтаж/ПНР:", finance_lines)

        if total_line:
            if left_lines and left_lines[-1].strip():
                left_lines.append("")
            left_lines.append(total_line)

        self.preview.setPlainText("\n".join(left_lines))
        self._set_spec_summary(spec_lines)
    except Exception as exc:
        self.preview.setPlainText(f"Не удалось прочитать данные: {exc}")
        self._set_spec_summary("Не удалось обновить сводку спецификаций.")
        try:
            self.refresh_spec_models(None)
        except Exception:
            pass


def _patch_page() -> None:
    """Apply GUI-only preview improvements after the core STULZ runtime loaded."""

    page_module = sys.modules.get("gui.pages.stulz_page")
    page_class = getattr(page_module, "StulzPage", None) if page_module is not None else None
    if page_class is None:
        return

    page_class._ensure_spec_summary_widget = _ensure_spec_summary_widget
    page_class._set_spec_summary = _set_spec_summary
    page_class.refresh_preview = refresh_preview

    app = QApplication.instance()
    if app is None:
        return

    for widget in app.allWidgets():
        if isinstance(widget, page_class):
            widget._ensure_spec_summary_widget()
            QTimer.singleShot(0, widget.refresh_preview)


_patch_page()
