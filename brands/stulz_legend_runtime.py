from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from PySide6.QtWidgets import QApplication, QCheckBox

import brands.stulz as _base
import brands.stulz_runtime as _runtime
import brands.stulz_ui_runtime as _ui
from core.excel_reader import CachedSheet, _find_quantity_row, parse_model_groups


# Preserve the complete public/runtime surface of the current STULZ module.
for _name in dir(_ui):
    if not _name.startswith("__"):
        globals()[_name] = getattr(_ui, _name)


_ORIGINAL_LOAD_CALC = _runtime.load_calc
_ORIGINAL_BUILD_DESCRIPTION = _runtime._build_offer_item_description
_ORIGINAL_BUILD_SPECIFICATION_BLOCKS = _runtime.build_specification_blocks


def _compact(value: object) -> str:
    text = str(value or "").lower().replace("ё", "е")
    return re.sub(r"[^a-zа-я0-9]+", "", text)


def _plain(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ")).strip()


def _number(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _plain(value).replace(" ", "").replace(",", ".").replace("%", "")
    if not text:
        return None
    try:
        return float(text)
    except Exception:
        return None


def _is_legend_candidate(value: object, model: str) -> bool:
    text = _plain(value)
    if not text:
        return False

    normalized = text.lower().replace("ё", "е")
    compact = _compact(text)
    if not compact or compact == _compact(model):
        return False

    if normalized in {
        "model",
        "модель",
        "q-ty",
        "qty",
        "quantity",
        "кол-во",
        "количество",
        "%",
    }:
        return False

    # Pure numbers in the header are normally quantities/revisions, not legends.
    if _number(text) is not None:
        return False

    return True


def _legend_for_group(
    sheet: CachedSheet,
    qty_col: int,
    amount_col: int,
    model: str,
    qty_row: int,
) -> str:
    """Read the free-text legend immediately above one model column group.

    Typical STULZ layout:
        D1 = Алматы
        D2 = ASR 552 AS
        C3 = quantity, D3+ = calculation values

    We also inspect the quantity-side column so merged headers spanning C:D work.
    """

    model_key = _compact(model)
    model_row: int | None = None
    upper_limit = min(sheet.max_row, max(12, qty_row))

    for row in range(1, upper_limit + 1):
        for col in (amount_col, qty_col):
            if _compact(sheet.cell(row, col).value) == model_key and model_key:
                model_row = row
                break
        if model_row is not None:
            break

    if model_row is None:
        model_row = max(2, qty_row - 1)

    # Same amount column first, then the quantity-side/adjacent cells to support
    # merged city/project labels. Stop after a small header window so we never
    # accidentally pick text from the calculation body.
    columns: list[int] = []
    for col in (amount_col, qty_col, amount_col - 1, qty_col - 1):
        if 1 <= col <= sheet.max_column and col not in columns:
            columns.append(col)

    for row in range(model_row - 1, max(0, model_row - 5), -1):
        for col in columns:
            value = sheet.cell(row, col).value
            if _is_legend_candidate(value, model):
                return _plain(value)

    return ""


def _read_item_legends(calc_path: Path, sheet_name: str | None, calc: Any) -> list[str]:
    """Return legends aligned with CalcData.items, preserving duplicate models."""

    legends = ["" for _item in getattr(calc, "items", [])]
    if not calc_path.exists() or not legends:
        return legends

    workbook = None
    try:
        workbook = load_workbook(calc_path, read_only=True, data_only=True)
        raw_sheet = (
            workbook[sheet_name]
            if sheet_name and sheet_name in workbook.sheetnames
            else workbook[workbook.sheetnames[0]]
        )
        sheet = CachedSheet(raw_sheet)
        groups = parse_model_groups(sheet)
        if not groups:
            return legends

        qty_row = _find_quantity_row(sheet, groups)
        records: list[tuple[str, str]] = []
        for qty_col, amount_col, model in groups:
            qty = _number(sheet.cell(qty_row, qty_col).value) or 0.0
            if qty <= 0:
                continue
            records.append(
                (
                    _compact(model),
                    _legend_for_group(sheet, qty_col, amount_col, model, qty_row),
                )
            )

        used: set[int] = set()
        for item_index, item in enumerate(calc.items):
            item_key = _compact(getattr(item, "name", ""))
            record_index = next(
                (
                    index
                    for index, (model_key, _legend) in enumerate(records)
                    if index not in used and model_key == item_key
                ),
                None,
            )
            if record_index is None:
                record_index = next((index for index in range(len(records)) if index not in used), None)
            if record_index is None:
                continue
            used.add(record_index)
            legends[item_index] = records[record_index][1]

        return legends
    except Exception:
        return legends
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass


def load_calc(context):
    """Load the normal STULZ calculation and attach header legends to its rows."""

    calc = _ORIGINAL_LOAD_CALC(context)
    legends = _read_item_legends(Path(context.calc_path), calc.sheet_name, calc)
    for item, legend in zip(calc.items, legends):
        # OfferItem is a regular dataclass (no slots), so runtime metadata can be
        # attached without changing the shared model used by other brands.
        setattr(item, "legend", legend)
    return calc


def _source_legend_hint(block: dict[str, Any]) -> str:
    """Extract a short physical-spec label when Excel does not provide one."""

    source_dir = Path(str(block.get("source_dir") or ""))
    folder = source_dir.name
    if folder:
        # Common supplier folder form:
        # SAM Trade LLP@KCELL_ASR552AS_Almaty@ASR 552 AS
        match = re.search(r"_([^_@]+)@[^@]+$", folder)
        if match:
            return _plain(match.group(1))

    label = _plain(block.get("source_label") or "")
    if label:
        return Path(label).name
    return ""


def _legends_for_spec_blocks(calc: Any, blocks: list[dict[str, Any]]) -> list[str]:
    """Align duplicate-model Excel legends with physical specification blocks.

    A source-folder text match is preferred when possible; otherwise duplicate
    models are paired in their original occurrence order. The physical folder is
    used only as a fallback when the Excel header has no usable legend.
    """

    records: list[tuple[int, str, str]] = []
    for index, item in enumerate(getattr(calc, "items", []) or []):
        records.append(
            (
                index,
                _compact(getattr(item, "name", "")),
                _plain(getattr(item, "legend", "")),
            )
        )

    used: set[int] = set()
    result: list[str] = []

    for block in blocks:
        model_key = _compact(block.get("model") or block.get("calc_model") or "")
        source_text = _compact(
            f"{block.get('source_label', '')} {block.get('source_dir', '')}"
        )

        candidates = [
            record
            for record in records
            if record[0] not in used and (not model_key or record[1] == model_key)
        ]

        chosen: tuple[int, str, str] | None = None
        for record in candidates:
            legend_key = _compact(record[2])
            if legend_key and source_text and legend_key in source_text:
                chosen = record
                break

        if chosen is None and candidates:
            chosen = candidates[0]

        if chosen is None:
            remaining = [record for record in records if record[0] not in used]
            if remaining:
                chosen = remaining[0]

        legend = ""
        if chosen is not None:
            used.add(chosen[0])
            legend = chosen[2]

        result.append(legend or _source_legend_hint(block))

    return result


def _spec_row_for_block(
    context: Any,
    block: dict[str, Any],
    block_index: int,
) -> dict[str, Any] | None:
    rows = [row for row in (getattr(context, "spec_models", []) or []) if row.get("enabled", True)]
    block_key = _plain(block.get("spec_key") or "").lower()
    if block_key:
        for row in rows:
            row_key = _plain(row.get("key") or "").lower()
            if row_key and row_key == block_key:
                return row

    if 0 <= block_index < len(rows):
        return rows[block_index]
    return None


def build_specification_blocks(context, calc):
    """Add the selected Excel legend to each physical STULZ specification block."""

    blocks, warnings = _ORIGINAL_BUILD_SPECIFICATION_BLOCKS(context, calc)
    legends = _legends_for_spec_blocks(calc, blocks)

    for index, (block, legend) in enumerate(zip(blocks, legends)):
        row = _spec_row_for_block(context, block, index)
        legend_enabled = bool(row.get("legend_enabled", True)) if row is not None else True
        block["legend"] = legend
        block["legend_enabled"] = legend_enabled

        if not legend_enabled or not legend:
            continue

        model = _plain(block.get("model") or block.get("calc_model") or "")
        suffix = f" — {legend}"
        block["options_title"] = f"Опции, включенные в комплектацию кондиционеров {model}{suffix}:"
        block["technical_specs_title"] = f"Технические характеристики кондиционеров {model}{suffix}:"

    return blocks, warnings


def _build_offer_item_description(item, block, options: dict[str, bool]) -> str:
    text = _ORIGINAL_BUILD_DESCRIPTION(item, block, options)
    legend = _plain(getattr(item, "legend", ""))
    if options.get("legend", True) and legend:
        return f"{legend} — {text}"
    return text


def _prepend_option(mapping: dict[str, Any], key: str, value: Any) -> dict[str, Any]:
    return {key: value, **{name: current for name, current in mapping.items() if name != key}}


def _install_legend_option_schema() -> None:
    """Teach the existing STULZ settings/UI about the new Legend checkbox."""

    _base.DESCRIPTION_OPTION_DEFAULTS = _prepend_option(
        dict(_base.DESCRIPTION_OPTION_DEFAULTS),
        "legend",
        True,
    )

    page_module = sys.modules.get("gui.pages.stulz_page")
    if page_module is not None and hasattr(page_module, "STULZ_DESCRIPTION_OPTION_DEFAULTS"):
        page_module.STULZ_DESCRIPTION_OPTION_DEFAULTS = _prepend_option(
            dict(page_module.STULZ_DESCRIPTION_OPTION_DEFAULTS),
            "legend",
            True,
        )

    page_runtime_module = sys.modules.get("gui.pages.stulz_page_runtime")
    if page_runtime_module is not None and hasattr(page_runtime_module, "DESCRIPTION_OPTION_LABELS"):
        page_runtime_module.DESCRIPTION_OPTION_LABELS = _prepend_option(
            dict(page_runtime_module.DESCRIPTION_OPTION_LABELS),
            "legend",
            "Легенда",
        )


def _ensure_legend_checkbox(widget: object) -> None:
    """Add Legend to an already-created description control group if necessary."""

    try:
        widget._ensure_description_options_controls()  # type: ignore[attr-defined]
        checkboxes = getattr(widget, "_stulz_description_checkboxes", {})
        if "legend" in checkboxes:
            return

        group = getattr(widget, "_stulz_description_group", None)
        layout = group.layout() if group is not None else None
        if layout is None:
            return

        checkbox = QCheckBox("Легенда")
        checkbox.setChecked(bool(widget.description_options().get("legend", True)))  # type: ignore[attr-defined]
        checkbox.toggled.connect(
            lambda _checked: widget._on_description_option_changed("legend")  # type: ignore[attr-defined]
        )
        checkboxes["legend"] = checkbox
        layout.insertWidget(0, checkbox)
    except Exception:
        return


# Patch all mature STULZ paths at module import time. The GUI preview calls the
# runtime module directly, while Word generation ultimately executes functions
# in brands.stulz, so both namespaces must point to the enriched implementations.
_install_legend_option_schema()
_base.load_calc = load_calc
_runtime.load_calc = load_calc
_ui.load_calc = load_calc
globals()["load_calc"] = load_calc

_base.build_specification_blocks = build_specification_blocks
_runtime.build_specification_blocks = build_specification_blocks
_ui.build_specification_blocks = build_specification_blocks
globals()["build_specification_blocks"] = build_specification_blocks

_base._build_offer_item_description = _build_offer_item_description
_runtime._build_offer_item_description = _build_offer_item_description
_ui._build_offer_item_description = _build_offer_item_description
globals()["_build_offer_item_description"] = _build_offer_item_description


app = QApplication.instance()
if app is not None:
    page_module = sys.modules.get("gui.pages.stulz_page")
    page_class = getattr(page_module, "StulzPage", None) if page_module is not None else None
    if page_class is not None:
        for widget in app.allWidgets():
            if isinstance(widget, page_class):
                _ensure_legend_checkbox(widget)
