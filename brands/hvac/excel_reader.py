from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover - user environment guard
    raise RuntimeError(
        "Для модуля HVAC нужна библиотека openpyxl. Добавьте openpyxl в requirements.txt."
    ) from exc


@dataclass(frozen=True)
class HVACPosition:
    key: str
    name: str
    qty: float | int | str | None
    amount: float | int | None
    source_col: int
    price_row_label: str

    @property
    def qty_text(self) -> str:
        return format_qty(self.qty)

    @property
    def amount_text(self) -> str:
        return format_money(self.amount)


def read_sheet_names(xlsx_path: str | Path) -> list[str]:
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel файл не найден: {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def read_hvac_positions(
    xlsx_path: str | Path,
    sheet_name: str | None = None,
    preferred_amount_rows: Iterable[str] = ("DDP Almaty", "TOTAL", "Total per quantity"),
    header_row: int = 2,
) -> list[HVACPosition]:
    """Read HVAC calculation columns as commercial offer positions.

    Expected calculation layout:
      row 2: Model / % / Q-ty / Duct ... / % / Q-ty / AHU ...
      column A: row labels, including Quantity and DDP Almaty/TOTAL.

    The function is intentionally tolerant: it finds columns where the header cell has
    equipment name and previous header cell is "Q-ty". Quantity is taken from the
    previous column on the Quantity row; amount is taken from the selected amount row
    in the equipment column.
    """
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel файл не найден: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

        quantity_row = _find_label_row(ws, ("Quantity", "QTY", "Qty", "Кол-во")) or 3
        amount_row, amount_label = _find_first_existing_label_row(ws, preferred_amount_rows)
        if not amount_row:
            raise ValueError(
                "Не нашёл строку суммы в Excel. Ожидал одну из строк: "
                + ", ".join(preferred_amount_rows)
            )

        positions: list[HVACPosition] = []
        for col in range(2, ws.max_column + 1):
            name = _clean_text(ws.cell(header_row, col).value)
            prev_header = _clean_text(ws.cell(header_row, col - 1).value)
            if not name or name in {"%", "Q-ty", "QTY", "Qty"}:
                continue
            if prev_header not in {"Q-ty", "QTY", "Qty", "Кол-во"}:
                continue

            qty = ws.cell(quantity_row, col - 1).value
            amount = _to_number(ws.cell(amount_row, col).value)
            if amount is None:
                continue
            if abs(float(amount)) < 0.000001:
                continue

            positions.append(
                HVACPosition(
                    key=_make_key(name, col),
                    name=name,
                    qty=qty,
                    amount=amount,
                    source_col=col,
                    price_row_label=amount_label,
                )
            )
        return positions
    finally:
        wb.close()


def _find_first_existing_label_row(ws, labels: Iterable[str]) -> tuple[int | None, str]:
    for label in labels:
        row = _find_label_row(ws, (label,))
        if row:
            return row, label
    return None, ""


def _find_label_row(ws, labels: Iterable[str]) -> int | None:
    normalized = {_normalize_label(x) for x in labels}
    for row in range(1, ws.max_row + 1):
        value = _normalize_label(ws.cell(row, 1).value)
        if value in normalized:
            return row
    return None


def _normalize_label(value) -> str:
    return _clean_text(value).casefold()


def _clean_text(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").split()).strip()


def _to_number(value) -> float | int | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return value
    try:
        return float(str(value).replace(" ", "").replace(",", "."))
    except ValueError:
        return None


def _make_key(name: str, col: int) -> str:
    safe = []
    for ch in name.lower():
        if ch.isalnum():
            safe.append(ch)
        elif safe and safe[-1] != "_":
            safe.append("_")
    key = "".join(safe).strip("_") or f"item_{col}"
    return f"{key}_{col}"


def format_qty(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def format_money(value) -> str:
    if value is None or value == "":
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    return f"{number:,.2f}".replace(",", " ")
