from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable
import re

from openpyxl import load_workbook


_QTY_ALIASES = ("quantity", "qty", "q-ty", "кол-во", "количество")
_MODEL_ALIASES = ("model", "модель", "equipment", "оборудование")
_UNIT_PRICE_ALIASES = (
    "total per unit",
    "price per unit",
    "unit price",
    "цена за единицу",
)
_TOTAL_PER_QTY_ALIASES = (
    "total per quantity",
    "total by quantity",
    "сумма по количеству",
)
_GRAND_TOTAL_ALIASES = ("grand total", "total", "итого")
_RATE_ALIASES = ("rate of currency", "exchange rate", "курс")
_VAT_ALIASES = ("vat, %", "vat %", "vat", "ндс, %", "ндс %", "ндс")
_ENGINEERING_ALIASES = ("engineering", "инжиниринг")
_INSTALLATION_ALIASES = ("installation", "монтаж")
_STARTUP_ALIASES = ("start-up", "startup", "commissioning", "пусконалад", "пуско-налад")
_DELIVERY_TERMS = ("DDP", "DAP", "EXW", "FCA", "CPT", "CIP")
_CURRENCY_LABELS = {
    "KZT": ("kzt", "тенге", "тг"),
    "EUR": ("eur", "euro", "€", "евро"),
    "USD": ("usd", "$", "доллар"),
    "RUB": ("rub", "руб", "₽"),
}


@dataclass(frozen=True)
class CalcItem:
    key: str
    name: str
    qty: float | int | str | None
    unit_price: float | int | None
    total_price: float | int | None
    source_col: int

    @property
    def qty_text(self) -> str:
        return format_qty(self.qty)

    @property
    def unit_price_text(self) -> str:
        return format_money(self.unit_price)

    @property
    def total_price_text(self) -> str:
        return format_money(self.total_price)

    # Compatibility with the first HVAC module.
    @property
    def amount(self) -> float | int | None:
        return self.total_price

    @property
    def amount_text(self) -> str:
        return self.total_price_text

    @property
    def price_row_label(self) -> str:
        return "TOTAL"


@dataclass(frozen=True)
class CalcService:
    included: bool | None
    percent: float | None = None
    amount: float | None = None
    source_label: str = ""


@dataclass
class CalcParseResult:
    file_path: Path
    sheet_name: str
    items: list[CalcItem] = field(default_factory=list)
    currency: str | None = None
    exchange_rate: float | None = None
    vat_percent: float | None = None
    vat_included: bool | None = None
    delivery_basis: str | None = None
    subtotal: float | None = None
    grand_total: float | None = None
    engineering: CalcService = field(default_factory=lambda: CalcService(None))
    installation: CalcService = field(default_factory=lambda: CalcService(None))
    startup: CalcService = field(default_factory=lambda: CalcService(None))
    warnings: list[str] = field(default_factory=list)


class SheetIndex:
    """Read a worksheet once and search it by normalized text markers."""

    def __init__(self, worksheet) -> None:
        self.title = worksheet.title
        value_rows: list[tuple[Any, ...]] = []
        format_rows: list[tuple[str, ...]] = []
        self.text_cells: list[tuple[int, int, str]] = []
        for row_no, cells in enumerate(worksheet.iter_rows(), start=1):
            values: list[Any] = []
            formats: list[str] = []
            for col_no, cell in enumerate(cells, start=1):
                value = cell.value
                values.append(value)
                formats.append(str(cell.number_format or ""))
                text = normalize_text(value)
                if text:
                    self.text_cells.append((row_no, col_no, text))
            value_rows.append(tuple(values))
            format_rows.append(tuple(formats))
        self.values = value_rows
        self.number_formats = format_rows
        self.max_row = len(self.values)
        self.max_column = max((len(row) for row in self.values), default=0)

    def value(self, row: int, col: int) -> Any:
        if row < 1 or col < 1 or row > self.max_row:
            return None
        values = self.values[row - 1]
        return values[col - 1] if col <= len(values) else None

    def number_format(self, row: int, col: int) -> str:
        if row < 1 or col < 1 or row > self.max_row:
            return ""
        formats = self.number_formats[row - 1]
        return formats[col - 1] if col <= len(formats) else ""

    def find(self, aliases: Iterable[str], *, exact: bool = False) -> list[tuple[int, int]]:
        normalized = [normalize_text(x) for x in aliases if normalize_text(x)]
        result: list[tuple[int, int]] = []
        for row, col, text in self.text_cells:
            matched = text in normalized if exact else any(alias in text for alias in normalized)
            if matched:
                result.append((row, col))
        return result


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").replace("ё", "е").casefold()
    text = re.sub(r"[\r\n\t_]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def read_sheet_names(file_path: str | Path) -> list[str]:
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel файл не найден: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def parse_calculation(
    file_path: str | Path,
    sheet_name: str | None = None,
    preferred_sheet: str | None = None,
) -> CalcParseResult:
    """Parse a SAM calculation by labels, without absolute row/column bindings."""

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel файл не найден: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        selected = _select_sheet(workbook, sheet_name or preferred_sheet)
        sheet = SheetIndex(workbook[selected])
        result = CalcParseResult(file_path=path, sheet_name=selected)

        qty_row = _find_quantity_row(sheet)
        model_row = _find_model_row(sheet, qty_row)
        unit_price_row = _find_last_row(sheet, _UNIT_PRICE_ALIASES)
        total_per_qty_row = _find_last_row(sheet, _TOTAL_PER_QTY_ALIASES)
        grand_total_row = _find_grand_total_row(sheet)

        if qty_row is None:
            result.warnings.append("Не найдена строка Quantity / Qty / Кол-во.")
        if model_row is None:
            result.warnings.append("Не найдена строка с наименованиями оборудования.")
        if unit_price_row is None:
            result.warnings.append("Не найдена строка Total per unit / Цена за единицу.")
        if total_per_qty_row is None and grand_total_row is None:
            result.warnings.append("Не найдена итоговая строка TOTAL / Total per quantity.")

        if qty_row and model_row:
            result.items = _extract_items(
                sheet,
                model_row=model_row,
                qty_row=qty_row,
                unit_price_row=unit_price_row,
                total_row=grand_total_row or total_per_qty_row,
            )

        result.exchange_rate = _find_numeric_near_label(sheet, _RATE_ALIASES)
        result.vat_percent = _find_numeric_near_label(sheet, _VAT_ALIASES)
        result.vat_included = _detect_vat_included(sheet, grand_total_row)
        result.currency = _detect_currency(sheet, result.exchange_rate, grand_total_row)
        result.delivery_basis = _detect_delivery_basis(sheet)
        result.grand_total = _find_row_total(sheet, grand_total_row)
        result.subtotal = _find_row_total(sheet, total_per_qty_row)
        result.engineering = _detect_service(sheet, _ENGINEERING_ALIASES, mode="engineering")
        result.installation = _detect_service(sheet, _INSTALLATION_ALIASES, mode="installation")
        result.startup = _detect_service(sheet, _STARTUP_ALIASES, mode="startup")

        if not result.items:
            result.warnings.append("Не найдено ни одной позиции с ненулевым количеством и суммой.")
        if result.currency is None:
            result.warnings.append("Не удалось определить валюту.")
        if result.delivery_basis is None:
            result.warnings.append("Не удалось определить условия поставки.")

        return result
    finally:
        workbook.close()


def _select_sheet(workbook, requested: str | None) -> str:
    if requested and requested in workbook.sheetnames:
        return requested
    if requested:
        wanted = normalize_text(requested)
        for name in workbook.sheetnames:
            if normalize_text(name) == wanted:
                return name

    best_name = workbook.sheetnames[0]
    best_score = -1
    for name in workbook.sheetnames:
        ws = workbook[name]
        score = 0
        title = normalize_text(name)
        if "ddp" in title:
            score += 5
        if "almaty" in title or "алматы" in title:
            score += 3
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), values_only=True):
            for value in row:
                text = normalize_text(value)
                if text in _QTY_ALIASES:
                    score += 5
                if text == "total":
                    score += 3
                if "total per unit" in text:
                    score += 2
        if score > best_score:
            best_name, best_score = name, score
    return best_name



def _find_quantity_row(sheet: SheetIndex) -> int | None:
    candidates = sheet.find(_QTY_ALIASES, exact=True)
    if not candidates:
        return None
    best_row = None
    best_score = -1
    for row, col in candidates:
        numeric_count = sum(
            1 for candidate_col in range(2, sheet.max_column + 1)
            if _to_number(sheet.value(row, candidate_col)) is not None
        )
        score = numeric_count * 10 + (5 if col == 1 else 0)
        if score > best_score:
            best_row, best_score = row, score
    return best_row

def _find_best_row(sheet: SheetIndex, aliases: Iterable[str], exact: bool = False) -> int | None:
    cells = sheet.find(aliases, exact=exact)
    return cells[0][0] if cells else None


def _find_last_row(sheet: SheetIndex, aliases: Iterable[str]) -> int | None:
    cells = sheet.find(aliases)
    return max((row for row, _ in cells), default=None)


def _find_model_row(sheet: SheetIndex, qty_row: int | None) -> int | None:
    cells = sheet.find(_MODEL_ALIASES, exact=True)
    if cells:
        return cells[0][0]
    if qty_row and qty_row > 1:
        return qty_row - 1
    return None


def _find_grand_total_row(sheet: SheetIndex) -> int | None:
    exact_total = sheet.find(("total", "итого"), exact=True)
    if exact_total:
        return max(row for row, _ in exact_total)
    return _find_last_row(sheet, _GRAND_TOTAL_ALIASES)


def _extract_items(
    sheet: SheetIndex,
    *,
    model_row: int,
    qty_row: int,
    unit_price_row: int | None,
    total_row: int | None,
) -> list[CalcItem]:
    items: list[CalcItem] = []
    for col in range(2, sheet.max_column + 1):
        name = _clean_display_text(sheet.value(model_row, col))
        if not name or _is_helper_header(name):
            continue

        qty = _to_number(sheet.value(qty_row, col))
        unit_price = _to_number(sheet.value(unit_price_row, col)) if unit_price_row else None
        total_price = _to_number(sheet.value(total_row, col)) if total_row else None

        # Do not turn technical summary columns into equipment positions.
        if qty in (None, 0) and total_price in (None, 0):
            continue
        if total_price in (None, 0) and unit_price not in (None, 0) and qty not in (None, 0):
            total_price = float(unit_price) * float(qty)
        if unit_price in (None, 0) and total_price not in (None, 0) and qty not in (None, 0):
            unit_price = float(total_price) / float(qty)

        items.append(
            CalcItem(
                key=_make_key(name, col),
                name=name,
                qty=qty,
                unit_price=unit_price,
                total_price=total_price,
                source_col=col,
            )
        )
    return items


def _is_helper_header(value: str) -> bool:
    text = normalize_text(value)
    return text in {"%", "q-ty", "qty", "quantity", "model", "модель"}


def _find_numeric_near_label(sheet: SheetIndex, aliases: Iterable[str]) -> float | None:
    for row, col in reversed(sheet.find(aliases)):
        for candidate_col in range(col + 1, min(sheet.max_column, col + 5) + 1):
            number = _to_number(sheet.value(row, candidate_col))
            if number is not None:
                return float(number)
        for candidate_row in range(row + 1, min(sheet.max_row, row + 3) + 1):
            number = _to_number(sheet.value(candidate_row, col))
            if number is not None:
                return float(number)
    return None


def _detect_currency(
    sheet: SheetIndex,
    exchange_rate: float | None,
    grand_total_row: int | None = None,
) -> str | None:
    # 1. Explicit labels in the sheet are the strongest signal.
    for currency, aliases in _CURRENCY_LABELS.items():
        if sheet.find(aliases, exact=True):
            return currency

    # 2. Inspect Excel financial/accounting number formats. The symbol may be
    # stored only in the cell style and therefore absent from the cell value.
    rows = [grand_total_row] if grand_total_row else []
    rows.extend(range(1, sheet.max_row + 1))
    seen_rows: set[int] = set()
    format_markers = {
        "KZT": ("₸", "kzt", "тенге", "тг", "[$₸"),
        "EUR": ("€", "eur", "euro", "евро", "[$€"),
        "USD": ("$", "usd", "доллар", "[$$"),
        "RUB": ("₽", "rub", "руб", "[$₽"),
    }
    scores = {currency: 0 for currency in format_markers}
    for row in rows:
        if not row or row in seen_rows:
            continue
        seen_rows.add(row)
        for col in range(1, sheet.max_column + 1):
            if _to_number(sheet.value(row, col)) is None:
                continue
            fmt = normalize_text(sheet.number_format(row, col))
            for currency, markers in format_markers.items():
                if any(normalize_text(marker) in fmt for marker in markers):
                    scores[currency] += 10 if row == grand_total_row else 1
    best_currency, best_score = max(scores.items(), key=lambda item: item[1])
    if best_score > 0:
        return best_currency

    # 3. A non-trivial exchange rate in SAM calculations normally means the
    # monetary rows are in KZT. This is only a last fallback.
    if exchange_rate and exchange_rate > 1.01:
        return "KZT"
    return None


def _detect_delivery_basis(sheet: SheetIndex) -> str | None:
    candidates: list[str] = []
    for row, col, text in sheet.text_cells:
        original = _clean_display_text(sheet.value(row, col))
        if any(re.search(rf"\b{term}\b", original, re.IGNORECASE) for term in _DELIVERY_TERMS):
            candidates.append(original)
    candidates.append(sheet.title)

    for preferred in ("DDP", "DAP", "FCA", "CPT", "CIP", "EXW"):
        for candidate in candidates:
            cleaned = _clean_delivery_basis(candidate)
            if cleaned and cleaned.upper().startswith(preferred) and "FOR US" not in cleaned.upper():
                return cleaned
    return None


def _clean_delivery_basis(value: str) -> str:
    text = value.replace("_", " ")
    text = re.sub(r"\s+", " ", text).strip()
    match = re.search(r"\b(DDP|DAP|EXW|FCA|CPT|CIP)\b\s*([^,;\n\r]*)", text, re.IGNORECASE)
    if not match:
        return ""
    term = match.group(1).upper()
    tail = re.sub(r"\([^)]*\)", "", match.group(2)).strip()
    tail = re.sub(r"\b\d{1,2}[-./]\d{1,2}(?:[-./]\d{2,4})?\b.*$", "", tail).strip(" -_/.,;")
    city = {"almaty": "Алматы", "алматы": "Алматы"}.get(normalize_text(tail), tail)
    return f"{term} {city}".strip()


def _find_row_total(sheet: SheetIndex, row: int | None) -> float | None:
    if row is None:
        return None
    values = [_to_number(sheet.value(row, col)) for col in range(2, sheet.max_column + 1)]
    numbers = [float(value) for value in values if value is not None]
    return numbers[-1] if numbers else None


def _detect_vat_included(sheet: SheetIndex, grand_total_row: int | None) -> bool | None:
    vat_rows = [row for row, _ in sheet.find(_VAT_ALIASES)]
    if not vat_rows:
        return None
    if grand_total_row is None:
        return any(_row_has_positive_number(sheet, row) for row in vat_rows)
    return any(row < grand_total_row and _row_has_positive_number(sheet, row) for row in vat_rows)


def _detect_service(sheet: SheetIndex, aliases: Iterable[str], *, mode: str) -> CalcService:
    matches = sheet.find(aliases)
    if not matches:
        # Installation may be represented as an equipment column header.
        if mode == "installation":
            for row, col, text in sheet.text_cells:
                if "монтаж" in text and "демонтаж" not in text:
                    amount = _max_positive_in_column(sheet, col)
                    return CalcService(amount > 0, amount=amount, source_label=_clean_display_text(sheet.value(row, col)))
        return CalcService(None)

    for row, col in reversed(matches):
        label = _clean_display_text(sheet.value(row, col))
        normalized = normalize_text(label)
        if mode == "installation" and "демонтаж" in normalized:
            continue
        if mode == "startup" and not any(alias in normalized for alias in ("start-up", "startup", "commissioning", "пуск")):
            continue

        percent = None
        if mode == "engineering":
            percent = _first_number_right(sheet, row, col)
        amount = _max_positive_in_row(sheet, row, start_col=col + 1)
        included = (percent is not None and percent > 0) or (amount is not None and amount > 0)
        if included or mode != "installation":
            return CalcService(included, percent=percent, amount=amount, source_label=label)

    if mode == "installation":
        for row, col, text in sheet.text_cells:
            if "монтаж" in text and "демонтаж" not in text and "пуск" not in text:
                amount = _max_positive_in_column(sheet, col)
                if amount is not None and amount > 0:
                    return CalcService(True, amount=amount, source_label=_clean_display_text(sheet.value(row, col)))
        return CalcService(False)

    return CalcService(None)


def _first_number_right(sheet: SheetIndex, row: int, col: int) -> float | None:
    for candidate_col in range(col + 1, min(sheet.max_column, col + 5) + 1):
        number = _to_number(sheet.value(row, candidate_col))
        if number is not None:
            return float(number)
    return None


def _max_positive_in_row(sheet: SheetIndex, row: int, start_col: int = 1) -> float | None:
    values = [_to_number(sheet.value(row, col)) for col in range(start_col, sheet.max_column + 1)]
    positives = [float(value) for value in values if value is not None and float(value) > 0]
    return max(positives) if positives else None


def _max_positive_in_column(sheet: SheetIndex, col: int) -> float | None:
    values = [_to_number(sheet.value(row, col)) for row in range(1, sheet.max_row + 1)]
    positives = [float(value) for value in values if value is not None and float(value) > 0]
    return max(positives) if positives else None


def _row_has_positive_number(sheet: SheetIndex, row: int) -> bool:
    return _max_positive_in_row(sheet, row) is not None


def _to_number(value: Any) -> float | int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return value
    text = str(value).replace("\xa0", " ").strip().replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _clean_display_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").replace("\n", " ").split()).strip()


def _make_key(name: str, col: int) -> str:
    safe = re.sub(r"[^\w]+", "_", name.casefold(), flags=re.UNICODE).strip("_")
    return f"{safe or 'item'}_{col}"


def format_qty(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace(".", ",")


def format_money(value: Any) -> str:
    if value is None or value == "":
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    return f"{number:,.2f}".replace(",", " ").replace(".", ",")


# First-iteration compatibility aliases for the existing HVAC module.
HVACPosition = CalcItem


def read_hvac_positions(
    xlsx_path: str | Path,
    sheet_name: str | None = None,
    **_: Any,
) -> list[CalcItem]:
    return parse_calculation(xlsx_path, sheet_name=sheet_name).items
