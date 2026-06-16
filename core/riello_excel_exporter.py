from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from core.riello_price import RielloPriceItem


@dataclass(frozen=True)
class RielloQuoteLine:
    item: RielloPriceItem
    qty: float
    note: str = ""

    @property
    def total(self) -> float:
        return float(self.item.price) * float(self.qty)

    @property
    def total_weight(self) -> float:
        return float(self.item.weight_kg) * float(self.qty)


@dataclass(frozen=True)
class RielloQuoteConfig:
    client_name: str
    city: str = "Алматы"
    currency: str = "EUR"
    ups_quantity: float = 1.0
    autonomy_min: str = ""
    battery_cabinet_type: str = ""
    rate: float = 1.0
    margin_percent: float = 15.0
    vat_percent: float = 0.0
    special_percent: float = 0.0
    transport_cost: float = 2000.0
    customs_clearance: float = 200.0
    certificate: float = 200.0
    transport_to_customer: float = 1500.0
    site_inspection: float = 0.0
    installation_startup: float = 0.0
    extra_cost: float = 0.0
    lines: list[RielloQuoteLine] = field(default_factory=list)

    @property
    def total_equipment(self) -> float:
        return sum(line.total for line in self.lines)

    @property
    def total_weight(self) -> float:
        return sum(line.total_weight for line in self.lines)


def _safe_float(value: Any, default: float = 0.0) -> float:
    text = str(value or "").strip().replace("\xa0", " ").replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except Exception:
        return default


def _fmt_qty(value: float) -> str:
    value = _safe_float(value)
    return str(int(value)) if value.is_integer() else str(value).replace(".", ",")


def _sanitize_filename(value: str) -> str:
    text = re.sub(r'[<>:"/\\|?*]+', "", str(value or "")).strip()
    text = re.sub(r"\s+", "_", text)
    return text or "Client"


def build_output_filename(config: RielloQuoteConfig, revision: int = 1, dt: datetime | None = None) -> str:
    dt = dt or datetime.now()
    first_model = config.lines[0].item.model if config.lines else "Riello"
    model = _sanitize_filename(first_model)
    client = _sanitize_filename(config.client_name)
    city = _sanitize_filename(config.city)
    return f"Riello_{model}_{client}_{city}_{dt:%d-%m-%y}_rev{revision}.xlsx"


def find_next_excel_revision(output_dir: Path, base_prefix: str) -> int:
    output_dir = Path(output_dir)
    max_rev = 0
    if output_dir.exists():
        for path in output_dir.glob("*.xlsx"):
            if path.name.startswith("~$"):
                continue
            if base_prefix and not path.stem.lower().startswith(base_prefix.lower()):
                continue
            match = re.search(r"rev\s*(\d+)", path.stem, re.IGNORECASE)
            if match:
                max_rev = max(max_rev, int(match.group(1)))
    return max_rev + 1 if max_rev else 1


def _first_sheet_starting(wb, prefix: str):
    prefix_lower = prefix.lower()
    for ws in wb.worksheets:
        if ws.title.lower().startswith(prefix_lower):
            return ws
    return wb.worksheets[0]


def _ensure_sheet(wb, name: str):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)


def _clear_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _copy_style(src, dst) -> None:
    if src is None or dst is None:
        return
    try:
        if src.has_style:
            dst._style = src._style.copy()
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = src.alignment.copy()
    except Exception:
        return


def _style_table_header(ws, row: int, min_col: int, max_col: int) -> None:
    fill = PatternFill("solid", fgColor="E5E7EB")
    side = Side(style="thin", color="CBD5E1")
    border = Border(left=side, right=side, top=side, bottom=side)
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row, col)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_table_body(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    side = Side(style="thin", color="E2E8F0")
    border = Border(left=side, right=side, top=side, bottom=side)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row, col)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_conf_sheet(wb, config: RielloQuoteConfig) -> int:
    ws = _ensure_sheet(wb, "conf")
    _clear_range(ws, 3, max(80, ws.max_row), 2, 8)
    start_row = 3
    for idx, line in enumerate(config.lines, start=start_row):
        ws.cell(idx, 2).value = line.item.model
        ws.cell(idx, 3).value = line.item.code
        ws.cell(idx, 4).value = line.item.price
        ws.cell(idx, 5).value = line.qty
        ws.cell(idx, 6).value = f"=E{idx}*D{idx}"
        ws.cell(idx, 7).value = line.item.weight_kg
        ws.cell(idx, 8).value = line.item.dimensions
    total_row = start_row + len(config.lines)
    ws.cell(total_row, 2).value = "TOTAL"
    ws.cell(total_row, 6).value = f"=SUM(F{start_row}:F{total_row - 1})" if config.lines else 0
    ws.cell(total_row, 7).value = f"=SUMPRODUCT(E{start_row}:E{total_row - 1},G{start_row}:G{total_row - 1})" if config.lines else 0
    ws.cell(total_row, 2).font = Font(bold=True)
    ws.cell(total_row, 6).font = Font(bold=True)
    ws.cell(total_row, 7).font = Font(bold=True)
    for col in (4, 6, 7):
        for row in range(start_row, total_row + 1):
            ws.cell(row, col).number_format = '#,##0.00'
    return total_row


def _write_ups_configuration(wb, config: RielloQuoteConfig) -> int:
    ws = _ensure_sheet(wb, "UPS configuration")
    _clear_range(ws, 1, max(80, ws.max_row), 1, 10)

    ws["A2"] = "V1"
    ws["C2"] = f"DDP {config.city} gross price without VAT"
    headers = ["Article", "Description", "Size", "Weight, kg", "per unit", "Quantity", f"TOTAL, {config.currency}", "Notes"]
    for offset, header in enumerate(headers, start=2):
        ws.cell(3, offset).value = header
    _style_table_header(ws, 3, 2, 9)

    start_row = 4
    for idx, line in enumerate(config.lines, start=start_row):
        item = line.item
        note_parts = []
        if item.power:
            note_parts.append(item.power)
        if line.note:
            note_parts.append(line.note)
        ws.cell(idx, 2).value = item.model
        ws.cell(idx, 3).value = item.code or item.description
        ws.cell(idx, 4).value = item.dimensions
        ws.cell(idx, 5).value = item.weight_kg
        ws.cell(idx, 6).value = item.price
        ws.cell(idx, 7).value = line.qty
        ws.cell(idx, 8).value = f"=G{idx}*F{idx}"
        ws.cell(idx, 9).value = "; ".join(note_parts)

    total_row = start_row + len(config.lines) + 1
    ws.cell(total_row, 6).value = f"Gross Total {config.currency}"
    ws.cell(total_row, 8).value = f"=SUM(H{start_row}:H{total_row - 2})" if config.lines else 0
    ws.cell(total_row, 6).font = Font(bold=True)
    ws.cell(total_row, 8).font = Font(bold=True)

    info_row = total_row + 2
    ws.cell(info_row, 2).value = "Время автономии"
    ws.cell(info_row, 3).value = config.autonomy_min
    ws.cell(info_row + 1, 2).value = "Тип батарейного шкафа"
    ws.cell(info_row + 1, 3).value = config.battery_cabinet_type
    ws.cell(info_row + 2, 2).value = "Суммарный вес, кг"
    ws.cell(info_row + 2, 3).value = config.total_weight

    _style_table_body(ws, start_row, max(start_row, total_row), 2, 9)
    widths = {"B": 22, "C": 24, "D": 24, "E": 12, "F": 14, "G": 12, "H": 16, "I": 32}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in range(start_row, total_row + 1):
        ws.row_dimensions[row].height = 30
    return total_row


def _write_ddp_sheet(wb, config: RielloQuoteConfig, conf_total_row: int) -> None:
    ws = _first_sheet_starting(wb, "DDP")
    first_model = config.lines[0].item.model if config.lines else "Riello UPS"
    first_code = config.lines[0].item.code if config.lines else ""

    ws["D2"] = first_model.replace(" ", "")
    ws["D3"] = first_code
    ws["C4"] = config.ups_quantity
    ws["D5"] = config.total_weight
    ws["C6"] = f"=conf!F{conf_total_row}"
    ws["D7"] = 1

    ws["A11"] = config.currency.lower()
    ws["D13"] = config.transport_cost
    ws["D14"] = config.customs_clearance
    ws["D15"] = config.extra_cost
    ws["D17"] = config.certificate
    ws["D19"] = config.transport_to_customer
    ws["D20"] = config.site_inspection
    ws["D21"] = config.installation_startup
    ws["B24"] = config.margin_percent
    ws["B26"] = config.vat_percent
    ws["A27"] = f"DDP {config.city}"
    ws["D29"] = config.rate
    ws["B31"] = config.vat_percent
    ws["B34"] = config.special_percent

    for cell in ("D5", "C6", "D8", "D13", "D14", "D15", "D17", "D19", "D20", "D21", "D23", "D24", "D25", "D26", "D27", "D29", "D30", "D31", "D32", "D34", "D35", "D36", "D37"):
        ws[cell].number_format = '#,##0.00'


def export_riello_excel(template_path: str | Path, output_path: str | Path, config: RielloQuoteConfig) -> Path:
    template_path = Path(template_path)
    output_path = Path(output_path)
    if not template_path.exists():
        raise FileNotFoundError(f"Excel-шаблон Riello не найден: {template_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(template_path, keep_links=False)
    conf_total_row = _write_conf_sheet(wb, config)
    _write_ups_configuration(wb, config)
    _write_ddp_sheet(wb, config, conf_total_row)

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    wb.save(output_path)
    return output_path
