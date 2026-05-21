from __future__ import annotations

from pathlib import Path
from typing import Any, Optional
import copy
import re

from openpyxl import load_workbook

from core.models import CalcData, OfferItem
from core.utils import as_float, first_not_empty


QTY_LABELS = {"q-ty", "qty", "quantity", "кол-во", "количество"}
TOTAL_LABELS = ("TOTAL", "ИТОГО")
TOTAL_PER_UNIT_LABELS = ("Total per unit", "Цена за единицу")
TOTAL_PER_QTY_LABELS = ("Total per quantity", "Сумма")
RATE_LABELS = ("Rate of currency", "Курс")
VAT_LABELS = ("VAT, %", "НДС, %", "VAT")
DELIVERY_LABELS = (
    "DDP", "DAP", "EXW", "FCA", "CPT", "CIP", "ALA", "Almaty", "Аксай", "Алматы", "Hamburg"
)
STOP_OPTION_LABELS = (
    "Total EXW for us", "Transport cost", "Customs clearance", "Certificate",
    "Storage", "Price without margin", "Margin", "Price + Margin", "VAT", "DDP", "DAP", "EXW",
)

_PARSE_CACHE: dict[tuple[str, int, int, str], CalcData] = {}
_SHEETS_CACHE: dict[tuple[str, int, int], list[str]] = {}


class _CellValue:
    __slots__ = ("row", "column", "value")

    def __init__(self, row: int, column: int, value: Any) -> None:
        self.row = row
        self.column = column
        self.value = value


class CachedSheet:
    """In-memory sheet snapshot.

    openpyxl is slow when the same sheet is scanned many times. This wrapper
    reads values once, then gives the existing parser cheap random access and
    cheap full-text indexes.
    """

    def __init__(self, ws) -> None:
        self.title = ws.title
        self._values = [tuple(row) for row in ws.iter_rows(values_only=True)]
        self.max_row = len(self._values)
        self.max_column = max((len(row) for row in self._values), default=0)
        self._norm_cells: list[tuple[int, int, str]] | None = None

    def cell(self, row: int, column: int) -> _CellValue:
        value = None
        if row >= 1 and column >= 1 and row <= self.max_row:
            row_values = self._values[row - 1]
            if column <= len(row_values):
                value = row_values[column - 1]
        return _CellValue(row, column, value)

    def __getitem__(self, address: str) -> _CellValue:
        from openpyxl.utils.cell import coordinate_to_tuple

        row, column = coordinate_to_tuple(address)
        return self.cell(row, column)

    def iter_rows(self):
        for row_idx, row_values in enumerate(self._values, start=1):
            yield [
                _CellValue(row_idx, col_idx, row_values[col_idx - 1] if col_idx <= len(row_values) else None)
                for col_idx in range(1, self.max_column + 1)
            ]

    def norm_cells(self) -> list[tuple[int, int, str]]:
        if self._norm_cells is None:
            cells: list[tuple[int, int, str]] = []
            for row_idx, row_values in enumerate(self._values, start=1):
                for col_idx, value in enumerate(row_values, start=1):
                    val = _norm(value)
                    if val:
                        cells.append((row_idx, col_idx, val))
            self._norm_cells = cells
        return self._norm_cells


def _file_signature(path: Path) -> tuple[str, int, int]:
    path = Path(path)
    stat = path.stat()
    return str(path.resolve()), stat.st_mtime_ns, stat.st_size


def clear_excel_cache() -> None:
    _PARSE_CACHE.clear()
    _SHEETS_CACHE.clear()


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def _norm(value: Any) -> str:
    s = _text(value).lower().replace("ё", "е")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _contains_any(value: Any, needles: tuple[str, ...] | list[str] | set[str]) -> bool:
    n = _norm(value)
    return any(_norm(x) in n for x in needles if _norm(x))


def _is_qty_label(value: Any) -> bool:
    n = _norm(value)
    return n in {_norm(x) for x in QTY_LABELS}


def find_cells_by_label(ws, labels: tuple[str, ...] | list[str] | set[str], exact: bool = False) -> list[tuple[int, int]]:
    result: list[tuple[int, int]] = []
    label_norms = [_norm(x) for x in labels]
    norm_cells = ws.norm_cells() if hasattr(ws, "norm_cells") else [
        (cell.row, cell.column, _norm(cell.value))
        for row in ws.iter_rows()
        for cell in row
        if _norm(cell.value)
    ]
    for row, col, val in norm_cells:
        if exact:
            if val in label_norms:
                result.append((row, col))
        else:
            if any(label in val for label in label_norms):
                result.append((row, col))
    return result


def find_row_by_label(ws, label: str) -> Optional[int]:
    cells = find_cells_by_label(ws, (label,), exact=True)
    return cells[0][0] if cells else None


def find_all_rows_by_label(ws, label: str) -> list[int]:
    return sorted({row for row, _col in find_cells_by_label(ws, (label,), exact=True)})


def _last_row_containing(ws, labels: tuple[str, ...]) -> Optional[int]:
    cells = find_cells_by_label(ws, labels, exact=False)
    return max((r for r, _c in cells), default=None)


def _first_row_containing(ws, labels: tuple[str, ...]) -> Optional[int]:
    cells = find_cells_by_label(ws, labels, exact=False)
    return min((r for r, _c in cells), default=None)


def detect_currency(exchange_rate: float, sheet_title: str, ws=None) -> str:
    title = _norm(sheet_title)
    if exchange_rate and exchange_rate > 1.01:
        return "KZT"
    if any(x in title for x in ("kzt", "тенг", "тг")):
        return "KZT"
    if ws is not None:
        for row in range(1, min(ws.max_row, 50) + 1):
            for col in range(1, min(ws.max_column, 20) + 1):
                val = _norm(ws.cell(row, col).value)
                if val in {"kzt", "тенге", "тг"}:
                    return "KZT" if exchange_rate and exchange_rate > 1.01 else "EUR"
                if val in {"eur", "euro", "€, euro", "евро"}:
                    return "EUR"
    return "EUR"


def _detect_delivery_basis(ws) -> str:
    title = _text(ws.title)
    candidates: list[str] = []
    if _contains_any(title, DELIVERY_LABELS):
        candidates.append(title)
    for row in range(1, min(ws.max_row, 80) + 1):
        val = _text(ws.cell(row, 1).value)
        if val and _contains_any(val, DELIVERY_LABELS):
            candidates.append(val)
    for val in candidates:
        n = _norm(val)
        if "ddp" in n:
            return val
        if "dap" in n:
            return val
        if "exw" in n:
            return val
    return "EXW Hamburg, Germany"


def _nearest_text_above(ws, row: int, col: int, depth: int = 3) -> str:
    for r in range(row - 1, max(0, row - depth - 1), -1):
        for c in (col, col - 1, col + 1):
            if c < 1 or c > ws.max_column:
                continue
            val = _text(ws.cell(r, c).value)
            if val and not _is_qty_label(val) and not val.replace(".", "", 1).isdigit():
                return val
    return ""


def _next_amount_col(ws, header_row: int, qty_col: int) -> int:
    # Usually Q-ty is immediately followed by a model/service amount column.
    # Skip another Q-ty marker if a template has an empty helper column.
    for col in range(qty_col + 1, min(ws.max_column, qty_col + 5) + 1):
        if not _is_qty_label(ws.cell(header_row, col).value):
            return col
    return qty_col + 1


def parse_model_groups(ws) -> list[tuple[int, int, str]]:
    """Find dynamic groups: (quantity column, amount column, item/model name)."""
    groups: list[tuple[int, int, str]] = []
    seen: set[tuple[int, int]] = set()

    # Search the full upper header area, not only row 2.
    for row in range(1, min(ws.max_row, 12) + 1):
        for col in range(1, ws.max_column + 1):
            if not _is_qty_label(ws.cell(row, col).value):
                continue
            amount_col = _next_amount_col(ws, row, col)
            if amount_col > ws.max_column:
                continue
            name = first_not_empty(
                ws.cell(row, amount_col).value,
                _nearest_text_above(ws, row, amount_col),
                ws.cell(row - 1, col).value if row > 1 else None,
            )
            if not name:
                continue
            name_s = _text(name)
            if _is_qty_label(name_s) or _norm(name_s) in {"model", "%"}:
                continue
            key = (col, amount_col)
            if key not in seen:
                groups.append((col, amount_col, name_s))
                seen.add(key)

    # Legacy fallback for older Stulz files.
    if not groups and first_not_empty(ws.cell(2, 4).value):
        groups.append((3, 4, _text(ws.cell(2, 4).value)))

    return groups


def _find_quantity_row(ws, groups: list[tuple[int, int, str]]) -> int:
    row = _first_row_containing(ws, ("Quantity", "Кол-во", "Количество"))
    if row:
        return row
    # Fallback: first row below header with numeric qty values in most groups.
    for r in range(1, min(ws.max_row, 20) + 1):
        hits = sum(1 for qty_col, _amount_col, _name in groups if as_float(ws.cell(r, qty_col).value, 0) > 0)
        if hits:
            return r
    return 4


def _value_from_row(ws, row: Optional[int], col: int, default: float = 0) -> float:
    if not row:
        return default
    return as_float(ws.cell(row, col).value, default)


def _best_total_rows(ws) -> tuple[Optional[int], Optional[int], Optional[int]]:
    total_row = _last_row_containing(ws, ("TOTAL", "ИТОГО"))
    total_per_unit_row = _last_row_containing(ws, ("Total per unit", "Цена за единицу"))
    total_per_qty_row = _last_row_containing(ws, ("Total per quantity", "Сумма"))
    return total_row, total_per_unit_row, total_per_qty_row


def _parse_calc_sheet(ws) -> CalcData:
    groups = parse_model_groups(ws)
    if not groups:
        raise ValueError(f"Не найдены колонки Q-ty/Qty на листе '{ws.title}'")

    version = _text(first_not_empty(ws["C1"].value, ws["A1"].value, "Version 1"))
    delivery_basis = _detect_delivery_basis(ws)
    qty_row = _find_quantity_row(ws, groups)
    total_row, total_per_unit_row, total_per_qty_row = _best_total_rows(ws)

    rate_cells = find_cells_by_label(ws, RATE_LABELS, exact=False)
    rate_row = rate_cells[-1][0] if rate_cells else None
    first_amount_col = groups[0][1]
    exchange_rate = _value_from_row(ws, rate_row, first_amount_col, 1) or 1

    vat_cells = find_cells_by_label(ws, VAT_LABELS, exact=False)
    vat_percent = 0.0
    if vat_cells:
        vr, vc = vat_cells[-1]
        for col in range(vc + 1, min(ws.max_column, vc + 8) + 1):
            vat_percent = as_float(ws.cell(vr, col).value, 0)
            if vat_percent:
                break

    currency = detect_currency(exchange_rate, ws.title, ws)

    items: list[OfferItem] = []
    for qty_col, amount_col, name in groups:
        qty = _value_from_row(ws, qty_row, qty_col, 0)
        if qty <= 0:
            continue

        total = _value_from_row(ws, total_row, amount_col, 0)
        unit = _value_from_row(ws, total_per_unit_row, amount_col, 0)
        if not total:
            total = _value_from_row(ws, total_per_qty_row, amount_col, 0)
        if not unit and total and qty:
            unit = total / qty
        if not total and unit and qty:
            total = unit * qty

        # Last fallback: take the largest numeric value in the amount column below the totals area.
        if not total:
            values = [as_float(ws.cell(r, amount_col).value, 0) for r in range(1, min(ws.max_row, 80) + 1)]
            values = [v for v in values if v > 0]
            total = max(values, default=0)
            unit = total / qty if qty else total

        if name and (qty > 0 or total > 0):
            items.append(OfferItem(len(items) + 1, name, qty, unit, total))

    if not items:
        raise ValueError(f"На листе '{ws.title}' колонки найдены, но позиции с количеством не найдены")

    options: list[tuple[str, float]] = []
    opt_row = _first_row_containing(ws, ("Options:", "Опции"))
    if opt_row:
        for row_idx in range(opt_row + 1, min(ws.max_row, opt_row + 120) + 1):
            name = _text(ws.cell(row_idx, 1).value)
            if not name:
                continue
            if _contains_any(name, STOP_OPTION_LABELS):
                break
            qty_value = sum(_value_from_row(ws, row_idx, qty_col, 0) for qty_col, _amount_col, _model in groups)
            if qty_value > 0:
                options.append((name, qty_value))

    return CalcData(
        sheet_name=ws.title,
        version=version,
        currency=currency,
        vat_percent=vat_percent,
        exchange_rate=exchange_rate,
        delivery_basis=delivery_basis,
        items=items,
        options=options,
    )


def _find_offer_header(ws) -> Optional[tuple[int, int, int, int]]:
    # Returns (header_row, name_col, qty_col, unit_col, total_col)
    for row in range(1, min(ws.max_row, 80) + 1):
        cols = {"name": None, "qty": None, "unit": None, "total": None}
        for col in range(1, ws.max_column + 1):
            n = _norm(ws.cell(row, col).value)
            if not n:
                continue
            if cols["name"] is None and any(x in n for x in ("наименование", "description", "model")):
                cols["name"] = col
            if cols["qty"] is None and any(x in n for x in ("кол-во", "quantity", "qty", "q-ty")):
                cols["qty"] = col
            if cols["unit"] is None and any(x in n for x in ("цена за единицу", "unit price", "price per unit")):
                cols["unit"] = col
            if cols["total"] is None and any(x in n for x in ("сумма", "total", "amount")):
                cols["total"] = col
        if all(cols.values()):
            return row, int(cols["name"]), int(cols["qty"]), int(cols["unit"]), int(cols["total"])
    return None


def _parse_offer_sheet(ws) -> CalcData:
    header = _find_offer_header(ws)
    if not header:
        raise ValueError(f"Не найдена таблица КП на листе '{ws.title}'")
    header_row, name_col, qty_col, unit_col, total_col = header

    items: list[OfferItem] = []
    for row in range(header_row + 1, ws.max_row + 1):
        name = _text(ws.cell(row, name_col).value)
        if not name:
            continue
        if _contains_any(name, ("ИТОГО", "GRAND TOTAL")):
            break
        if _contains_any(name, ("ВСЕГО", "SUBTOTAL")):
            continue
        qty = as_float(ws.cell(row, qty_col).value, 0)
        unit = as_float(ws.cell(row, unit_col).value, 0)
        total = as_float(ws.cell(row, total_col).value, 0)
        if qty > 0 and (unit or total):
            if not unit and total:
                unit = total / qty
            if not total and unit:
                total = unit * qty
            items.append(OfferItem(len(items) + 1, name, qty, unit, total))

    if not items:
        raise ValueError(f"Таблица КП на листе '{ws.title}' найдена, но позиции не прочитаны")

    currency = "EUR"
    for row in range(1, min(ws.max_row, 80) + 1):
        for col in range(1, ws.max_column + 1):
            n = _norm(ws.cell(row, col).value)
            if "kzt" in n or "тенг" in n:
                currency = "KZT"
            if "eur" in n or "евро" in n:
                currency = "EUR"

    return CalcData(
        sheet_name=ws.title,
        version=_text(first_not_empty(ws.cell(1, 1).value, "Version 1")),
        currency=currency,
        vat_percent=0,
        exchange_rate=1,
        delivery_basis="DDP",
        items=items,
        options=[],
    )


def _sheet_priority(name: str, selected: Optional[str] = None) -> int:
    n = _norm(name)
    if selected and name == selected:
        return -100
    if n.startswith("кп") or "offer" in n or "commercial" in n:
        return 20
    if "spare" in n:
        return 10
    return 0


def parse_stulz_calc(xlsx_path: Path, sheet_name: Optional[str] = None) -> CalcData:
    path = Path(xlsx_path)
    signature = _file_signature(path)
    cache_key = (*signature, sheet_name or "")
    if cache_key in _PARSE_CACHE:
        return copy.deepcopy(_PARSE_CACHE[cache_key])

    wb = load_workbook(path, read_only=True, data_only=True)

    sheet_names = [sheet_name] if sheet_name and sheet_name in wb.sheetnames else list(wb.sheetnames)
    sheet_names = sorted(sheet_names, key=lambda s: _sheet_priority(s, sheet_name))

    errors: list[str] = []

    # 1) Try calculation sheets first. Each sheet is materialized once into
    # CachedSheet, and all parser lookups use that in-memory snapshot/index.
    for name in sheet_names:
        ws = CachedSheet(wb[name])
        try:
            if _find_offer_header(ws) and (_norm(name).startswith("кп") or "offer" in _norm(name)):
                continue
            result = _parse_calc_sheet(ws)
            _PARSE_CACHE[cache_key] = copy.deepcopy(result)
            return result
        except Exception as exc:
            errors.append(f"{name}: {exc}")

    # 2) Fallback to ready КП/offer sheets.
    offer_candidates = sorted(wb.sheetnames, key=lambda s: _sheet_priority(s, sheet_name), reverse=True)
    for name in offer_candidates:
        ws = CachedSheet(wb[name])
        try:
            result = _parse_offer_sheet(ws)
            _PARSE_CACHE[cache_key] = copy.deepcopy(result)
            return result
        except Exception as exc:
            errors.append(f"{name} КП: {exc}")

    raise ValueError("Не удалось прочитать Excel. Подробности: " + " | ".join(errors[:6]))


def list_sheets(xlsx_path: Path) -> list[str]:
    path = Path(xlsx_path)
    signature = _file_signature(path)
    if signature in _SHEETS_CACHE:
        return list(_SHEETS_CACHE[signature])
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = list(wb.sheetnames)
    _SHEETS_CACHE[signature] = list(sheets)
    return sheets


def read_calc_excel(xlsx_path: Path, sheet_name: Optional[str] = None) -> CalcData:
    """Backward-compatible wrapper used by older brand modules."""
    return parse_stulz_calc(Path(xlsx_path), sheet_name)
